import sqlite3
import json
import hashlib
from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
import io
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- INICIALIZAÇÃO DO APP ---
app = Flask(__name__)
CORS(app)

CORS(app, resources={r"/api/*": {"origins": ["https://site.suporteverde.com.br/md/", "http://localhost:5173"]}})

DB_NAME = "sistema_notas_v2.db"

# --- 1. CONFIGURAÇÃO DO BANCO DE DADOS ---
def init_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    # Cria a tabela com ID Auto Incremental e Hash Único
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS notas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            empresa TEXT,
            data TEXT,
            valor TEXT,
            status TEXT,
            is_cadastrado INTEGER, 
            arquivo_base64 TEXT,
            detalhes_json TEXT,
            hash_registro TEXT UNIQUE
        )
    ''')
    conn.commit()
    conn.close()

# Inicializa o banco ao rodar o script
init_db()

# --- FUNÇÕES AUXILIARES ---
def limpar_moeda(valor):
    if isinstance(valor, (int, float)):
        return valor
    try:
        return float(str(valor).replace('R$', '').replace('.', '').replace(',', '.').strip())
    except:
        return 0.0
    
# Gera um ID único baseado no conteúdo para evitar duplicatas
def gerar_hash(empresa, data, valor):
    raw = f"{empresa}-{data}-{valor}"
    return hashlib.md5(raw.encode('utf-8')).hexdigest()

# --- 2. ROTAS DE BANCO DE DADOS ---

# Rota para LISTAR todas as notas salvas (GET)
@app.route('/api/notas', methods=['GET'])
def get_notas():
    try:
        conn = sqlite3.connect(DB_NAME)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM notas ORDER BY id DESC") # Mais recentes primeiro
        rows = cursor.fetchall()
        
        notas = []
        for row in rows:
            notas.append({
                "id": row["id"], # ID Incremental do SQLite
                "empresa": row["empresa"],
                "data": row["data"],
                "valor": row["valor"],
                "status": row["status"],
                "isCadastrado": bool(row["is_cadastrado"]),
                "arquivoBase64": row["arquivo_base64"],
                "detalhesCompletos": json.loads(row["detalhes_json"]) if row["detalhes_json"] else {}
            })
        conn.close()
        return jsonify(notas)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Rota para SALVAR novas notas emitidas (POST)
@app.route('/api/notas', methods=['POST'])
def save_notas():
    try:
        novas_notas = request.json
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()

        salvas = 0
        ignoradas = 0

        for nota in novas_notas:
            # Gera o hash para verificação de duplicidade
            rec_hash = gerar_hash(nota['empresa'], nota['data'], nota['valor'])
            
            # Verifica se já existe
            cursor.execute("SELECT id FROM notas WHERE hash_registro = ?", (rec_hash,))
            if cursor.fetchone() is None:
                cursor.execute('''
                    INSERT INTO notas (empresa, data, valor, status, is_cadastrado, arquivo_base64, detalhes_json, hash_registro)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    nota['empresa'], nota['data'], nota['valor'], nota['status'],
                    1 if nota['isCadastrado'] else 0,
                    nota['arquivoBase64'],
                    json.dumps(nota['detalhesCompletos']),
                    rec_hash
                ))
                salvas += 1
            else:
                ignoradas += 1
        
        conn.commit()
        conn.close()
        
        msg = f"{salvas} notas salvas."
        if ignoradas > 0:
            msg += f" ({ignoradas} duplicatas já existiam)"

        return jsonify({"message": msg, "duplicates": ignoradas}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Rota para ATUALIZAR status de cadastro (PUT)
@app.route('/api/notas/<id>', methods=['PUT'])
def update_nota(id):
    try:
        dados = request.json
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute('UPDATE notas SET is_cadastrado = ? WHERE id = ?', (1 if dados['isCadastrado'] else 0, id))
        conn.commit()
        conn.close()
        return jsonify({"message": "Status atualizado!"})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# --- 3. ROTA DE PROCESSAMENTO (ATUALIZADA COM LÓGICA DE DATA) ---
@app.route('/api/processar-notas', methods=['POST'])
def processar_notas():
    try:
        if 'file' not in request.files:
            return {'error': 'Nenhum arquivo enviado'}, 400
        
        file = request.files['file']
        
        # --- NOVOS CAMPOS PARA DATA ---
        modo_data = request.form.get('modoData', 'atual') # 'atual', 'venda', 'escolher'
        data_custom = request.form.get('dataCustom', '')

        # Leitura do arquivo
        if file.filename.endswith('.csv'):
            try:
                df = pd.read_csv(file, encoding='utf-8')
            except:
                file.seek(0)
                df = pd.read_csv(file, encoding='latin1', sep=';')
        else:
            df = pd.read_excel(file)

        df.columns = df.columns.str.strip()
        
        resultado_processamento = []

        # Estilos Excel
        estilo_titulo = Font(bold=True, size=14, color="FFFFFF")
        fundo_azul = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
        estilo_negrito = Font(bold=True)
        borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alinhamento_centro = Alignment(horizontal='center', vertical='center')

        for index, row in df.iterrows():
            wb = Workbook()
            ws = wb.active
            ws.title = "Nota Fiscal"

            # --- DEFINIÇÃO DA DATA DE EMISSÃO ---
            data_emissao = datetime.now().strftime("%d/%m/%Y") # Padrão
            
            if modo_data == 'escolher' and data_custom:
                # Converte YYYY-MM-DD para DD/MM/YYYY
                try:
                    data_obj = datetime.strptime(data_custom, '%Y-%m-%d')
                    data_emissao = data_obj.strftime("%d/%m/%Y")
                except:
                    pass
            elif modo_data == 'venda':
                # Tenta pegar da planilha
                if 'Data' in df.columns and pd.notna(row['Data']):
                    data_raw = row['Data']
                    # Se for objeto datetime do pandas/excel
                    if isinstance(data_raw, datetime):
                        data_emissao = data_raw.strftime("%d/%m/%Y")
                    else:
                        data_emissao = str(data_raw)

            # Lógica de Busca de Nome
            nome_cliente = 'Consumidor'
            buscas = ['Resp. Fin', 'Resp Fin', 'Resp. Fin.', 'Nome', 'Cliente', 'Razão Social']
            for b in buscas:
                if b in df.columns and pd.notna(row.get(b)):
                    nome_cliente = row.get(b)
                    break
            
            if nome_cliente == 'Consumidor':
                for col in df.columns:
                    limpa = col.lower().replace('.', '').replace(' ', '')
                    if limpa in ['respfin', 'nome', 'cliente', 'razaosocial']:
                        if pd.notna(row.get(col)):
                            nome_cliente = row.get(col)
                            break
            
            # Dados para o Frontend
            val_devido = limpar_moeda(row.get('V. Devido', 0))
            
            item_dados = {
                'temp_id': index,  # Usamos temp_id pois ainda não foi pro banco
                'nome_arquivo': f"NF-{1000+index} - {str(nome_cliente)[:30]}",
                'respFin': nome_cliente,
                'origem': row.get('Origem', '-'),
                'cpf': row.get('CPF/CNPJ') or row.get('CPF') or '-',
                'titulo': row.get('Título', 'Serviço'),
                'especie': row.get('Espécie', 'NF-e'),
                'vDevido': f"R$ {val_devido:,.2f}".replace('.', ','),
                'vReceb': f"R$ {limpar_moeda(row.get('V. Receb', 0)):,.2f}".replace('.', ','),
                'vDesc': f"R$ {limpar_moeda(row.get('V. Desc', 0)):,.2f}".replace('.', ','),
                'pContas': row.get('P. Contas', 'Fidelizado'),
                'cpfResp': row.get('CPF Resp', row.get('CPF/CNPJ', '-')),
                'data': data_emissao, # Usa a data calculada acima
                'venc': str(row.get('Venc', datetime.now().strftime("%d/%m/%Y"))),
                'arquivo': '' 
            }

            # Montagem do Excel (Visual)
            ws.merge_cells('A1:D2')
            cell = ws['A1']
            cell.value = "MD SISTEMAS - NOTA FISCAL DE SERVIÇO"
            cell.font = estilo_titulo
            cell.fill = fundo_azul
            cell.alignment = alinhamento_centro

            ws['A4'] = "Número da Nota:"
            ws['B4'] = 1000 + index
            ws['C4'] = "Data Emissão:"
            ws['D4'] = data_emissao # Data dinâmica no Excel
            ws['A4'].font = ws['C4'].font = estilo_negrito

            ws.merge_cells('A6:D6')
            ws['A6'] = "DADOS DO TOMADOR DE SERVIÇO"
            ws['A6'].font = Font(bold=True, color="2C5282")
            ws['A6'].border = Border(bottom=Side(style='thick', color="2C5282"))

            ws['A8'] = "Razão Social / Nome:"
            ws['B8'] = nome_cliente
            ws['A9'] = "CPF / CNPJ:"
            ws['B9'] = item_dados['cpf']
            ws['A10'] = "Origem:"
            ws['B10'] = item_dados['origem']

            ws.merge_cells('A12:D12')
            ws['A12'] = "DETALHES DO PAGAMENTO"
            ws['A12'].font = Font(bold=True, color="2C5282")
            ws['A12'].border = Border(bottom=Side(style='thick', color="2C5282"))

            headers = ["Descrição (Espécie)", "Vencimento", "Desconto", "Valor Total"]
            col_letters = ['A', 'B', 'C', 'D']
            for i, header in enumerate(headers):
                cell = ws[f'{col_letters[i]}14']
                cell.value = header
                cell.font = estilo_negrito
                cell.border = borda_fina
                cell.alignment = alinhamento_centro

            val_desc = limpar_moeda(row.get('V. Desc', 0))
            descricao = f"{row.get('Espécie', 'Serviço')} - {row.get('Título', '')}"
            
            ws['A15'] = descricao
            ws['B15'] = item_dados['venc']
            ws['C15'] = val_desc
            ws['D15'] = val_devido

            ws['C15'].number_format = 'R$ #,##0.00'
            ws['D15'].number_format = 'R$ #,##0.00'

            for col in col_letters:
                ws[f'{col}15'].border = borda_fina
                ws[f'{col}15'].alignment = Alignment(horizontal='center')

            ws['C17'] = "VALOR LÍQUIDO:"
            ws['C17'].font = estilo_negrito
            ws['D17'] = val_devido - val_desc
            ws['D17'].number_format = 'R$ #,##0.00'
            ws['D17'].font = Font(bold=True, size=12)

            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            arquivo_base64 = base64.b64encode(excel_buffer.getvalue()).decode('utf-8')
            item_dados['arquivo'] = arquivo_base64
            
            resultado_processamento.append(item_dados)

        return jsonify(resultado_processamento)

    except Exception as e:
        print(f"Erro detalhado: {e}")
        return {'error': str(e)}, 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)